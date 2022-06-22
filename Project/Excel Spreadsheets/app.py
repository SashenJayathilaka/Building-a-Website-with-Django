import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')

sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1 , 1)

# print(cell.value)
# print(sheet.max_row)

for row in range(2 , sheet.max_row + 1):
    cell = sheet.cell(row ,3)
    # print(cell.value)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price


wb.save('transactions2.xlsx')